import asyncio
import logging
import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from concurrent.futures import ThreadPoolExecutor
import functools
import matplotlib.pyplot as plt
import io

from aiogram.types import FSInputFile
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile,
    ReplyKeyboardRemove, InputFile, BufferedInputFile
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# --- Настройки ---
BOT_TOKEN = "8717360750:AAEbHfHYt0wk0JtwbCvuJxBFK08-lSxqitM"
ADMIN_IDS = [341440758, 885305710]

EXPERT_IDS = [341440758, 885305710]

EXCEL_FILE = "cdlqi_results.xlsx"
HISTORY_FILE = "user_history.json"  # Файл для хранения истории тестов
ACHIEVEMENTS_FILE = "user_achievements.json"  # Файл для хранения достижений

logging.basicConfig(level=logging.INFO)

bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

executor = ThreadPoolExecutor(max_workers=4)

# --- Хранилища данных ---
user_answers: Dict[int, Dict[int, int]] = {}
user_history: Dict[int, List[Dict]] = {}  # История тестов пользователя
user_achievements: Dict[int, List[str]] = {}  # Достижения пользователя

# --- Загрузка и сохранение данных ---
def load_json_data(filename: str, default: dict) -> dict:
    """Загружает данные из JSON файла"""
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return default
    return default

def save_json_data(filename: str, data: dict):
    """Сохраняет данные в JSON файл"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# Загружаем историю и достижения при старте
user_history = load_json_data(HISTORY_FILE, {})
user_history = {int(k): v for k, v in user_history.items()}

user_achievements = load_json_data(ACHIEVEMENTS_FILE, {})
user_achievements = {int(k): v for k, v in user_achievements.items()}

# --- FSM состояния ---
class CDLQIForm(StatesGroup):
    waiting_for_answer = State()

class ExpertConsultationForm(StatesGroup):
    waiting_for_skin_type = State()
    waiting_for_problems = State()
    waiting_for_budget = State()
    waiting_for_additional = State()

# --- Список вопросов ---
QUESTIONS = [
    "Вопрос 1/10: У моего подростка болит кожа?",
    "Вопрос 2/10: Состояние кожи моего подростка влияет на качество его сна?",
    "Вопрос 3/10: Мой подросток беспокоится, что его кожное заболевание может быть серьезным?",
    "Вопрос 4/10: Состояние кожи моего подростка затрудняет посещение школы или занятия спортом?",
    "Вопрос 5/10: Состояние кожи моего подростка затрудняет общение с друзьями и другими детьми его возраста?",
    "Вопрос 6/10: Состояние кожи моего подростка вызывает у него грусть?",
    "Вопрос 7/10: Состояние кожи моего подростка вызывает жжение или покалывание?",
    "Вопрос 8/10: Мой подросток склонен оставаться дома из-за своего кожного заболевания?",
    "Вопрос 9/10: Мой подросток беспокоится о том, что у него останутся шрамы от кожного заболевания?",
    "Вопрос 10/10: Кожа моего подростка зудит?"
]

# --- Функции для работы с Excel ---
def init_excel_file():
    """Создает Excel файл с заголовками, если он не существует."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "CDLQI Results"
        headers = [
            "ID пользователя", "Имя пользователя", "Дата и время",
            "Вопрос 1", "Вопрос 2", "Вопрос 3", "Вопрос 4", "Вопрос 5",
            "Вопрос 6", "Вопрос 7", "Вопрос 8", "Вопрос 9", "Вопрос 10",
            "Общий балл", "Уровень влияния"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(EXCEL_FILE)

def _save_results_to_excel_sync(user_id: int, username: str, answers: List[int], total_score: int, impact_level: str):
    """Синхронная функция сохранения результатов в Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        init_excel_file()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    row = [
        user_id,
        username or "Нет username",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        *answers,
        total_score,
        impact_level
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)
    logging.info(f"✅ Результаты пользователя {user_id} сохранены в Excel")

async def save_results_to_excel_async(user_id: int, username: str, answers: List[int], total_score: int, impact_level: str):
    """Асинхронная обёртка для сохранения в Excel"""
    loop = asyncio.get_event_loop()
    func = functools.partial(
        _save_results_to_excel_sync,
        user_id, username, answers, total_score, impact_level
    )
    await loop.run_in_executor(executor, func)

# --- Функции для истории тестов ---
def save_test_to_history(user_id: int, total_score: int, impact_level: str):
    """Сохраняет результат теста в историю пользователя"""
    global user_history
    
    user_id_str = str(user_id)
    if user_id_str not in user_history:
        user_history[user_id_str] = []
    
    user_history[user_id_str].append({
        "date": datetime.now().isoformat(),
        "score": total_score,
        "impact": impact_level
    })
    
    # Сохраняем в файл
    save_json_data(HISTORY_FILE, user_history)

def get_user_history(user_id: int) -> List[Dict]:
    """Возвращает историю тестов пользователя"""
    return user_history.get(str(user_id), [])

# --- Функции для достижений ---
def check_and_award_achievements(user_id: int, total_score: int) -> List[str]:
    """Проверяет и начисляет достижения"""
    global user_achievements
    
    user_id_str = str(user_id)
    if user_id_str not in user_achievements:
        user_achievements[user_id_str] = []
    
    new_achievements = []
    history = get_user_history(user_id)
    
    # Достижение: Первый тест
    if len(history) == 1 and "first_test" not in user_achievements[user_id_str]:
        user_achievements[user_id_str].append("first_test")
        new_achievements.append("🎯 Первый шаг — пройден первый тест!")
    
    # Достижение: Улучшение
    if len(history) >= 2:
        last_two = history[-2:]
        if last_two[1]["score"] < last_two[0]["score"]:
            if "improvement" not in user_achievements[user_id_str]:
                user_achievements[user_id_str].append("improvement")
                new_achievements.append("📈 На пути к лучшему — результат улучшился!")
    
    # Достижение: Значительное улучшение (на 5+ баллов)
    if len(history) >= 2:
        first = history[0]["score"]
        last = history[-1]["score"]
        if first - last >= 5:
            if "big_improvement" not in user_achievements[user_id_str]:
                user_achievements[user_id_str].append("big_improvement")
                new_achievements.append("🌟 Крутой прогресс — минус 5+ баллов!")
    
    # Достижение: 3 теста
    if len(history) >= 3 and "three_tests" not in user_achievements[user_id_str]:
        user_achievements[user_id_str].append("three_tests")
        new_achievements.append("🎓 Исследователь — пройдено 3 теста!")
    
    # Достижение: Минимальное влияние
    if total_score <= 5 and "low_impact" not in user_achievements[user_id_str]:
        user_achievements[user_id_str].append("low_impact")
        new_achievements.append("🍃 Чистая кожа — минимальное влияние акне!")
    
    # Сохраняем достижения
    save_json_data(ACHIEVEMENTS_FILE, user_achievements)
    
    return new_achievements

def get_achievement_badge(achievement_id: str) -> str:
    """Возвращает эмодзи-значок для достижения"""
    badges = {
        "first_test": "🎯",
        "improvement": "📈",
        "big_improvement": "🌟",
        "three_tests": "🎓",
        "low_impact": "🍃",
        "expert_consult": "👩‍⚕️"
    }
    return badges.get(achievement_id, "🏆")

# --- Функция для создания графика ---
async def create_progress_chart(user_id: int) -> Optional[BufferedInputFile]:
    """Создает график прогресса пользователя"""
    history = get_user_history(user_id)
    
    if len(history) < 2:
        return None
    
    # Подготавливаем данные
    dates = [datetime.fromisoformat(h["date"]) for h in history]
    scores = [h["score"] for h in history]
    
    # Создаем график
    plt.figure(figsize=(10, 6))
    plt.plot(dates, scores, marker='o', linewidth=2, markersize=8, color='#4CAF50')
    plt.fill_between(dates, scores, alpha=0.3, color='#4CAF50')
    
    plt.title('📊 Динамика CDLQI', fontsize=16, pad=20)
    plt.xlabel('Дата', fontsize=12)
    plt.ylabel('Баллы CDLQI', fontsize=12)
    plt.grid(True, alpha=0.3)
    
    # Добавляем цветовые зоны
    plt.axhspan(0, 1, alpha=0.1, color='green', label='Минимальное')
    plt.axhspan(2, 5, alpha=0.1, color='lightgreen', label='Лёгкое')
    plt.axhspan(6, 10, alpha=0.1, color='yellow', label='Умеренное')
    plt.axhspan(11, 20, alpha=0.1, color='orange', label='Значительное')
    plt.axhspan(21, 40, alpha=0.1, color='red', label='Экстремальное')
    
    plt.legend(loc='upper right')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    # Сохраняем в буфер
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    buf.seek(0)
    plt.close()
    
    return BufferedInputFile(buf.read(), filename="progress.png")

# --- Клавиатуры ---
def get_answer_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура с оценками от 0 до 4."""
    buttons = [
        [KeyboardButton(text="0️⃣ Никогда"), KeyboardButton(text="1️⃣ Редко")],
        [KeyboardButton(text="2️⃣ Иногда"), KeyboardButton(text="3️⃣ Часто")],
        [KeyboardButton(text="4️⃣ Всегда")]
    ]
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

def get_main_menu_keyboard() -> InlineKeyboardMarkup:
    """Главное меню."""
    buttons = [
        [InlineKeyboardButton(text="🔍 CDLQI-тест", callback_data="start_test")],
        [InlineKeyboardButton(text="📊 Моя статистика", callback_data="my_stats")],
        [InlineKeyboardButton(text="🏆 Мои достижения", callback_data="my_achievements")],
        [InlineKeyboardButton(text="👩‍⚕️ Консультация эксперта", callback_data="expert_consult")],
        [InlineKeyboardButton(text="📅 7-дневный план", callback_data="seven_day_plan")],
        [InlineKeyboardButton(text="💬 Чат родителей", callback_data="parent_chat")],
        [InlineKeyboardButton(text="🚨 Срочно к врачу?", callback_data="urgent")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def get_seven_day_plan_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура для 7-дневного плана."""
    buttons = [
        [InlineKeyboardButton(text="💎 День 1", callback_data="day_1"),
         InlineKeyboardButton(text="💎 День 2", callback_data="day_2"),
         InlineKeyboardButton(text="💎 День 3", callback_data="day_3")],
        [InlineKeyboardButton(text="💎 День 4", callback_data="day_4"),
         InlineKeyboardButton(text="💎 День 5", callback_data="day_5"),
         InlineKeyboardButton(text="💎 День 6", callback_data="day_6")],
        [InlineKeyboardButton(text="🎉 День 7", callback_data="day_7")],
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def get_skin_type_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура для выбора типа кожи"""
    buttons = [
        [KeyboardButton(text="🧴 Жирная"), KeyboardButton(text="💧 Сухая")],
        [KeyboardButton(text="🔄 Комбинированная"), KeyboardButton(text="😊 Нормальная")],
        [KeyboardButton(text="❓ Не знаю")]
    ]
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

def get_budget_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура для выбора бюджета"""
    buttons = [
        [KeyboardButton(text="💰 До 1000₽"), KeyboardButton(text="💰💰 1000-3000₽")],
        [KeyboardButton(text="💰💰💰 3000-5000₽"), KeyboardButton(text="💎 Любой")],
        [KeyboardButton(text="🔙 Отмена")]
    ]
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

# --- Обработчики команд ---
@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    """Обработчик команды /start с картинкой"""
    await state.clear()
    if message.from_user.id in user_answers:
        del user_answers[message.from_user.id]

    try:
        photo = FSInputFile("1.PNG")
    except:
        photo = None
    
    welcome_text = (
        "👋 <b>Привет! Я бот NeGovoriProidet — помогаю родителям поддержать подростков с акне!</b> 💙\n\n"
        "✨ <b>Новые возможности:</b>\n"
        "• 📊 История тестов и график прогресса\n"
        "• 🏆 Достижения за активность\n"
        "• 👩‍⚕️ Консультация реального врача дерматолога\n\n"
        "Выберите, что хотите сделать:"
    )
    
    if photo:
        await message.answer_photo(
            photo=photo,
            caption=welcome_text,
            reply_markup=get_main_menu_keyboard(),
            parse_mode="HTML"
        )
    else:
        await message.answer(
            welcome_text,
            reply_markup=get_main_menu_keyboard(),
            parse_mode="HTML"
        )

@dp.message(Command("result"))
async def cmd_result(message: Message):
    """Обработчик команды /result. Отправляет Excel файл администраторам."""
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("⛔ У вас нет прав для выполнения этой команды.")
        return

    if not os.path.exists(EXCEL_FILE):
        await message.answer("📊 База данных результатов пока пуста.")
        return

    file = FSInputFile(EXCEL_FILE)
    await message.answer_document(file, caption="📊 База результатов CDLQI-тестов.")

# --- Обработчики статистики и достижений ---
@dp.callback_query(F.data == "my_stats")
async def callback_my_stats(callback: CallbackQuery):
    """Показывает статистику пользователя"""
    await callback.answer()
    await callback.message.delete()
    
    user_id = callback.from_user.id
    history = get_user_history(user_id)
    
    if not history:
        await callback.message.answer(
            "📊 У вас пока нет пройденных тестов.\n\n"
            "Пройдите первый тест, чтобы увидеть статистику!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 Пройти тест", callback_data="start_test")],
                [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
            ])
        )
        return
    
    # Основная статистика
    last_test = history[-1]
    first_test = history[0]
    total_tests = len(history)
    
    if total_tests >= 2:
        improvement = first_test["score"] - last_test["score"]
        trend = f"⬇️ Улучшение на {improvement} баллов" if improvement > 0 else \
                f"⬆️ Ухудшение на {abs(improvement)} баллов" if improvement < 0 else \
                "⏸️ Без изменений"
    else:
        trend = "📝 Нужно больше тестов для динамики"
    
    stats_text = (
        f"📊 <b>Ваша статистика</b>\n\n"
        f"📝 Всего тестов: {total_tests}\n"
        f"🆕 Последний результат: {last_test['score']} баллов\n"
        f"📈 {trend}\n\n"
        f"📅 Первый тест: {datetime.fromisoformat(first_test['date']).strftime('%d.%m.%Y')}\n"
        f"🎯 Последний тест: {datetime.fromisoformat(last_test['date']).strftime('%d.%m.%Y')}"
    )
    
    # Пробуем создать график
    chart = await create_progress_chart(user_id)
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🏆 Мои достижения", callback_data="my_achievements")],
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ])
    
    if chart:
        await callback.message.answer_photo(
            photo=chart,
            caption=stats_text,
            reply_markup=keyboard,
            parse_mode="HTML"
        )
    else:
        await callback.message.answer(
            stats_text,
            reply_markup=keyboard,
            parse_mode="HTML"
        )

@dp.callback_query(F.data == "my_achievements")
async def callback_my_achievements(callback: CallbackQuery):
    """Показывает достижения пользователя"""
    await callback.answer()
    await callback.message.delete()
    
    user_id = callback.from_user.id
    user_id_str = str(user_id)
    
    achievements_text = "🏆 <b>Ваши достижения</b>\n\n"
    
    if user_id_str in user_achievements and user_achievements[user_id_str]:
        for ach in user_achievements[user_id_str]:
            badge = get_achievement_badge(ach)
            if ach == "first_test":
                achievements_text += f"{badge} <b>Первый шаг</b> — пройден первый тест\n"
            elif ach == "improvement":
                achievements_text += f"{badge} <b>На пути к лучшему</b> — результат улучшился\n"
            elif ach == "big_improvement":
                achievements_text += f"{badge} <b>Крутой прогресс</b> — минус 5+ баллов\n"
            elif ach == "three_tests":
                achievements_text += f"{badge} <b>Исследователь</b> — пройдено 3 теста\n"
            elif ach == "low_impact":
                achievements_text += f"{badge} <b>Чистая кожа</b> — минимальное влияние акне\n"
            elif ach == "expert_consult":
                achievements_text += f"{badge} <b>Профессионал</b> — получена консультация эксперта\n"
    else:
        achievements_text += "У вас пока нет достижений. Проходите тесты и получайте награды! ✨"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Моя статистика", callback_data="my_stats")],
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(
        achievements_text,
        reply_markup=keyboard,
        parse_mode="HTML"
    )

# --- Обработчики консультации эксперта ---
@dp.callback_query(F.data == "expert_consult")
async def callback_expert_consult(callback: CallbackQuery, state: FSMContext):
    """Начинает процесс консультации с экспертом"""
    await callback.answer()
    await callback.message.delete()
    
    await state.set_state(ExpertConsultationForm.waiting_for_skin_type)
    
    text = (
        "👩‍⚕️ <b>Консультация врача-дерматолога</b>\n\n"
        "Ответьте на несколько вопросов, и наш эксперт подберёт "
        "персональные рекомендации по уходу.\n\n"
        "1/4: <b>Какой тип кожи у вашего подростка?</b>"
    )
    
    await callback.message.answer(
        text,
        reply_markup=get_skin_type_keyboard(),
        parse_mode="HTML"
    )

@dp.message(ExpertConsultationForm.waiting_for_skin_type)
async def process_skin_type(message: Message, state: FSMContext):
    """Обрабатывает тип кожи"""
    skin_type = message.text.replace("🧴 ", "").replace("💧 ", "").replace("🔄 ", "").replace("😊 ", "")
    
    if skin_type == "🔙 Отмена":
        await state.clear()
        await message.answer(
            "Консультация отменена. Возвращайтесь, когда будете готовы!",
            reply_markup=get_main_menu_keyboard()
        )
        return
    
    await state.update_data(skin_type=skin_type)
    await state.set_state(ExpertConsultationForm.waiting_for_problems)
    
    await message.answer(
        "2/4: <b>Опишите основные проблемы кожи</b> (например: прыщи, чёрные точки, жирный блеск, покраснения)\n\n"
        "Можно перечислить через запятую:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="HTML"
    )

@dp.message(ExpertConsultationForm.waiting_for_problems)
async def process_problems(message: Message, state: FSMContext):
    """Обрабатывает описание проблем"""
    problems = message.text
    
    await state.update_data(problems=problems)
    await state.set_state(ExpertConsultationForm.waiting_for_budget)
    
    await message.answer(
        "3/4: <b>Какой бюджет на уход в месяц?</b>",
        reply_markup=get_budget_keyboard(),
        parse_mode="HTML"
    )

@dp.message(ExpertConsultationForm.waiting_for_budget)
async def process_budget(message: Message, state: FSMContext):
    """Обрабатывает бюджет"""
    budget = message.text
    
    if budget == "🔙 Отмена":
        await state.clear()
        await message.answer(
            "Консультация отменена. Возвращайтесь, когда будете готовы!",
            reply_markup=get_main_menu_keyboard()
        )
        return
    
    await state.update_data(budget=budget)
    await state.set_state(ExpertConsultationForm.waiting_for_additional)
    
    await message.answer(
        "4/4: <b>Дополнительная информация</b>\n\n"
        "Напишите, если хотите что-то добавить (возраст подростка, аллергии, "
        "используемые средства и т.д.) или отправьте \"Нет\", если всё:\n\n"
        "<i>Например: Возраст 15 лет, используем гель для умывания, аллергии нет</i>",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="HTML"
    )

@dp.message(ExpertConsultationForm.waiting_for_additional)
async def process_additional(message: Message, state: FSMContext):
    """Завершает сбор информации и отправляет эксперту"""
    additional = message.text
    data = await state.get_data()
    user = message.from_user
    
    # Формируем заявку для эксперта
    consult_text = (
        "👩‍⚕️ <b>НОВАЯ ЗАЯВКА НА КОНСУЛЬТАЦИЮ</b>\n\n"
        f"👤 Пользователь: @{user.username or 'нет'} (ID: {user.id})\n"
        f"📝 Имя: {user.full_name}\n\n"
        f"🧴 <b>Тип кожи:</b> {data['skin_type']}\n"
        f"⚠️ <b>Проблемы:</b> {data['problems']}\n"
        f"💰 <b>Бюджет:</b> {data['budget']}\n"
        f"📋 <b>Дополнительно:</b> {additional}\n\n"
        f"📅 Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    
    # Отправляем всем экспертам
    success_count = 0
    errors = []
    
    for expert_id in EXPERT_IDS:
        try:
            await bot.send_message(expert_id, consult_text, parse_mode="HTML")
            success_count += 1
        except Exception as e:
            errors.append(f"Ошибка отправки эксперту {expert_id}: {e}")
            logging.error(f"Ошибка при отправке эксперту {expert_id}: {e}")
    
    # Если хотя бы одному эксперту отправилось успешно
    if success_count > 0:
        # Добавляем достижение за консультацию
        user_id = user.id
        user_id_str = str(user_id)
        if user_id_str not in user_achievements:
            user_achievements[user_id_str] = []
        if "expert_consult" not in user_achievements[user_id_str]:
            user_achievements[user_id_str].append("expert_consult")
            save_json_data(ACHIEVEMENTS_FILE, user_achievements)
        
        # Сообщение пользователю
        expert_word = "экспертам" if len(EXPERT_IDS) > 1 else "эксперту"
        await message.answer(
            f"✅ <b>Заявка отправлена!</b>\n\n"
            f"Наши {expert_word} получили вашу информацию и свяжутся с вами в ближайшее время "
            f"(обычно в течение 24 часов).\n\n"
            f"✨ Вы получили достижение <b>«Профессионал»</b> за обращение к эксперту!",
            reply_markup=get_main_menu_keyboard(),
            parse_mode="HTML"
        )
    else:
        error_text = "\n".join(errors[:3])  # Показываем первые 3 ошибки
        await message.answer(
            f"❌ Произошла ошибка при отправке заявки.\n"
            f"Попробуйте позже или свяжитесь с поддержкой.\n\n"
            f"<i>Техническая информация: {error_text}</i>",
            reply_markup=get_main_menu_keyboard(),
            parse_mode="HTML"
        )
    
    await state.clear()

# --- Обработчики колбэков (остальные) ---
@dp.callback_query(F.data == "start_test")
async def callback_start_test(callback: CallbackQuery, state: FSMContext):
    """Начинает тест CDLQI."""
    await callback.answer()
    await callback.message.delete()

    user_answers[callback.from_user.id] = {}
    await state.set_state(CDLQIForm.waiting_for_answer)
    await state.update_data(current_q=0)

    await callback.message.answer(
        f"🔍 <b>CDLQI-тест</b>\n\n{QUESTIONS[0]}",
        reply_markup=get_answer_keyboard(),
        parse_mode="HTML"
    )

@dp.callback_query(F.data == "back_to_main")
async def callback_back_to_main(callback: CallbackQuery, state: FSMContext):
    """Возвращает на главную."""
    await callback.answer()
    await state.clear()
    await callback.message.delete()
    await callback.message.answer(
        "👋 <b>Главное меню</b>",
        reply_markup=get_main_menu_keyboard(),
        parse_mode="HTML"
    )

# --- Обработчики 7-дневного плана ---
@dp.callback_query(F.data == "seven_day_plan")
async def callback_seven_day_plan(callback: CallbackQuery):
    """Показывает меню 7-дневного плана"""
    await callback.answer()
    await callback.message.delete()
    
    await callback.message.answer(
        "📅 <b>7-дневный план поддержки Bioderma</b>\n\nВыберите день:",
        reply_markup=get_seven_day_plan_keyboard(),
        parse_mode="HTML"
    )

@dp.callback_query(F.data.startswith("day_"))
async def callback_show_day(callback: CallbackQuery):
    """Показывает описание выбранного дня"""
    await callback.answer()
    await callback.message.delete()
    
    day_num = callback.data.split("_")[1]
    
    days_info = {
        "1": "💎 <b>День 1/7: Эмпатия и первый шаг</b>\n\n🗣️ Скажи: «Вижу, как тебе тяжело...»\n🧴 Уход: Sébium Gel Moussant\n📸 Домашка: Сфоткай кожу «до»",
        "2": "💎 <b>День 2/7: Ритуал ухода + позитив</b>\n\n🌅 Утро: Sébium Gel Moussant + Thermal Water\n🌃 Вечер: Sébium Lotion",
        "3": "💎 <b>День 3/7: Развеиваем мифы и триггеры</b>\n\n❌ Миф: «Сладкое вызывает акне» — нет\n➕ Добавь Sébium Global утром",
        "4": "💎 <b>День 4/7: Интенсивный уход и фотофиксация</b>\n\n🧴 Полный ритуал: Gel Moussant + Global + Cicabio\n📸 Сравни фото",
        "5": "💎 <b>День 5/7: Мотивация и закрепление</b>\n\n🗣️ Скажи: «Твоя кожа улучшается!»\n🧴 Полный уход + Thermal Water",
        "6": "💎 <b>День 6/7: Сообщество и дневник</b>\n\n💬 Чат родителей: «CDLQI 15→8»\n📸 Дневник: загрузи фото",
        "7": "🎉 <b>День 7/7: Итог и следующий шаг</b>\n\n📸 Сравни фото, перепройди тест\n⚠️ Если CDLQI >10 — к врачу"
    }
    
    text = days_info.get(day_num, "Информация готовится...")
    back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 К списку дней", callback_data="seven_day_plan")],
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(text, reply_markup=back_keyboard, parse_mode="HTML")

@dp.callback_query(F.data == "parent_chat")
async def callback_parent_chat(callback: CallbackQuery):
    """Раздел родительского чата."""
    await callback.answer()
    await callback.message.delete()
    
    text = (
        "💬 <b>Реальные родители делятся историями:</b>\n\n"
        "👤 «CDLQI 18→8: Sébium Global помог!»\n"
        "👤 «Cicabio спас от шрамов»\n\n"
        "Поделись анонимно своей историей или присоединяйся к чату."
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 Перейти в чат", url="https://t.me/+ws3h1RkmlnkxOTZi")],
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(text, reply_markup=keyboard, parse_mode="HTML")

@dp.callback_query(F.data == "urgent")
async def callback_urgent(callback: CallbackQuery):
    """Срочный раздел."""
    await callback.answer()
    await callback.message.delete()
    
    text = (
        "🚨 <b>Срочно к врачу, если:</b>\n\n"
        "⚠️ CDLQI >20\n"
        "⚠️ Шрамы, кровотечение, гной\n"
        "⚠️ Депрессия, изоляция\n"
        "⚠️ Нет улучшения за 4 недели\n\n"
        "Подготовь фото, результаты теста и триггеры.\n"
        "Найди врача и запишись на консультацию."
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 На главную", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(text, reply_markup=keyboard, parse_mode="HTML")

# --- Обработчик для FSM (ответы на вопросы теста) ---
@dp.message(CDLQIForm.waiting_for_answer)
async def process_answer(message: Message, state: FSMContext):
    """Обрабатывает ответы пользователя на вопросы теста."""
    user_id = message.from_user.id
    answer_text = message.text

    score_map = {
        "0️⃣ Никогда": 0, "1️⃣ Редко": 1, "2️⃣ Иногда": 2,
        "3️⃣ Часто": 3, "4️⃣ Всегда": 4
    }

    if answer_text not in score_map:
        await message.answer("Пожалуйста, используйте кнопки для ответа.")
        return

    score = score_map[answer_text]

    data = await state.get_data()
    current_q_index = data.get('current_q', 0)

    if user_id not in user_answers:
        user_answers[user_id] = {}
    user_answers[user_id][current_q_index] = score

    next_q_index = current_q_index + 1

    if next_q_index < len(QUESTIONS):
        await state.update_data(current_q=next_q_index)
        await message.answer(
            f"🔍 <b>CDLQI-тест</b>\n\n{QUESTIONS[next_q_index]}",
            reply_markup=get_answer_keyboard(),
            parse_mode="HTML"
        )
    else:
        await state.clear()

        all_answers = [user_answers[user_id][i] for i in range(len(QUESTIONS))]
        total_score = sum(all_answers)

        if total_score <= 1:
            impact = "Минимальное — акне не мешает жизни."
        elif total_score <= 5:
            impact = "Лёгкое — комедоны, лёгкое воспаление."
        elif total_score <= 10:
            impact = "Умеренное — видимые высыпания."
        elif total_score <= 20:
            impact = "Значительное — влияет на учёбу и общение."
        else:
            impact = "Экстремальное — депрессия, шрамы, срочно к дерматологу!"

        if total_score <= 5:
            recommendation = "Лёгкая акне: Sébium Gel Moussant + Sébium Lotion. Набор «Старт» 890₽."
        elif total_score <= 10:
            recommendation = "Средняя акне I: Sébium Global + Thermal Spring Water. Дуо «Контроль» 1450₽."
        elif total_score <= 20:
            recommendation = "Средняя акне II: Sébium Kerato+ + Cicabio Crème. Комплект «Интенсив» 2100₽."
        else:
            recommendation = "Тяжёлая акне: Cicabio Arnica+ + Ретинол Booster + консультация врача. SOS-набор 1850₽."

        username = message.from_user.username or message.from_user.full_name
        
        # Сохраняем в Excel асинхронно
        asyncio.create_task(
            save_results_to_excel_async(user_id, username, all_answers, total_score, impact)
        )
        
        # Сохраняем в историю
        save_test_to_history(user_id, total_score, impact)
        
        # Проверяем достижения
        new_achievements = check_and_award_achievements(user_id, total_score)

        result_text = (
            f"✅ <b>Твой CDLQI: {total_score}/40</b>\n\n"
            f"<b>Влияние акне на жизнь подростка:</b>\n{impact}\n\n"
            f"<b>💎 Рекомендация Bioderma:</b>\n{recommendation}"
        )

        await message.answer(result_text, reply_markup=ReplyKeyboardRemove(), parse_mode="HTML")
        
        # Показываем новые достижения
        if new_achievements:
            achievements_text = "🎉 <b>Новые достижения!</b>\n\n" + "\n".join([f"• {ach}" for ach in new_achievements])
            await message.answer(achievements_text, parse_mode="HTML")
        
        del user_answers[user_id]

        await message.answer(
            "Выберите дальнейшее действие:",
            reply_markup=get_main_menu_keyboard()
        )

# --- Запуск бота ---
async def main():
    init_excel_file()
    
    # Проверяем наличие matplotlib для графиков
    try:
        import matplotlib
        logging.info("✅ Matplotlib установлен, графики будут работать")
    except ImportError:
        logging.warning("⚠️ Matplotlib не установлен. Установите: pip install matplotlib")
    
    logging.info("🤖 Бот запущен и готов к работе!")
    await dp.start_polling(bot, skip_updates=True)

if __name__ == "__main__":
    asyncio.run(main())